from openpyxl import load_workbook, drawing
import pandas as pd
from xlrd import open_workbook
from tkinter.messagebox import showerror, showwarning, showinfo
import PIL
import json


class ExpenseLog :

    def __init__(self):

        self.foodPrice = dict()
        self.lineNumber = 0
        self.personNumber = 0

    def foodData(self, dej, din, soup, milV, milCV, prov):
        
        self.foodPrice["Déjeuner"] = dej
        self.foodPrice["Dîner"] = din
        self.foodPrice["Souper"] = soup
        self.foodPrice["V"] = milV
        self.foodPrice["CV"] = milCV
        self.foodPrice["Prov"] = prov
        
    def processExpLog(self, log_Rea, depense_Paquet):

        numOccur = []
        months = ['janvier', 'février', 'mars','avril','mai','juin','juillet','août','sept','octobre','novembre','décembre']
       
        wb = load_workbook(depense_Paquet)

        content = pd.read_excel(log_Rea, skiprows=5)
        data = content.head(1000)
        data = data.drop([data.columns[0] , data.columns[1]] ,  axis='columns')
        headers = [ x for x in list(data.columns.values) if 'Unnamed' not in x]
        s= 13
        
        for head in range(len(headers)):
            try:

                travLogR = data[data[headers[head]] == 'R']
                
                for index, row in travLogR.iterrows() :
                    s = 13
                    
                    if row[headers[head]] == 'R':
                        depenseSheets = wb.sheetnames
                        if row[headers[0]] not in depenseSheets:
                            sheetLocal = wb['Temp']
                            localOne = wb.copy_worksheet(sheetLocal)
                            localOne.title = row[headers[0]]
                            imgs = drawing.image.Image('logo.png')
                            localOne.add_image(imgs)
                            localOne.sheet_view.zoomScale = 65
                            #wb.save(depense_Paquet)
                        self.lineNumber += 1
                        sheet= wb[row[headers[0]]]
                        j,m,a = headers[head].split()
                        monthNum = months.index(m)
                        
                        if row[headers[0]] in numOccur :
                            occurName = numOccur.count(row[headers[0]])
                            s += occurName
                        else:
                            s = 13
                        
                        sheet.cell(row=4,column=11).value = row[headers[0]]
                        
                        sheet.cell(row=s,column=2).value = int(j)
                        sheet.cell(row=s,column=3).value = monthNum + 1
                        sheet.cell(row=s,column=4).value = int(a)
                        sheet.cell(row=s,column=6).value = self.foodPrice["Prov"]
                        sheet.cell(row=s,column=7).value = 'Repas (per diem)'

                        sheet.cell(row=s,column=14).value = int(self.foodPrice["Déjeuner"])
                        sheet.cell(row=s,column=15).value = int(self.foodPrice["Dîner"])
                        sheet.cell(row=s,column=16).value = int(self.foodPrice["Souper"])

                        numOccur.append(row[headers[0]])

                    #wb.save(depense_Paquet)

            except :


                Warning(data[headers[head]])

            try :

                travLogV = data[data[headers[head]] == 'V']

                for index, row in travLogV.iterrows() :
                    s = 13
                
                    if row[headers[head]] == 'V':
                        depenseSheets = wb.sheetnames
                        if row[headers[0]] not in depenseSheets:
                            sheetLocal = wb['Temp']
                            localOne = wb.copy_worksheet(sheetLocal)
                            localOne.title = row[headers[0]]
                            imgs = drawing.image.Image('logo.png')
                            localOne.add_image(imgs)
                            localOne.sheet_view.zoomScale = 65
                            #wb.save(depense_Paquet)
                        self.lineNumber += 1
                        sheet= wb[row[headers[0]]]
                        j,m,a = headers[head].split()
                        monthNum = months.index(m)

                        if row[headers[0]] in numOccur :
                            occurName = numOccur.count(row[headers[0]])
                            s += occurName
                        else:
                            s = 13

                        sheet.cell(row=4,column=11).value = row[headers[0]]

                        sheet.cell(row=s,column=2).value = int(j)
                        sheet.cell(row=s,column=3).value = monthNum + 1
                        sheet.cell(row=s,column=4).value = int(a)
                        sheet.cell(row=s,column=6).value = self.foodPrice["Prov"]
                        sheet.cell(row=s,column=7).value = 'Mileage'

                        sheet.cell(row=s,column=14).value = int(self.foodPrice["Déjeuner"])
                        sheet.cell(row=s,column=15).value = int(self.foodPrice["Dîner"])
                        sheet.cell(row=s,column=16).value = int(self.foodPrice["Souper"])

                        sheet.cell(row=s,column=10).value = int(self.foodPrice["V"])

                        numOccur.append(row[headers[0]])
                    # wb.save(depense_Paquet)

            except :

                Warning(data[headers[head]])


            try :
                
                travLogCV = data[data[headers[head]] == 'CV']

                for index, row in travLogCV.iterrows() :
                    s = 13

                    if row[headers[head]] == 'CV':
                        depenseSheets = wb.sheetnames

                        if row[headers[0]] not in depenseSheets:
                            sheetLocal = wb['Temp']
                            localOne = wb.copy_worksheet(sheetLocal)
                            localOne.title = row[headers[0]]
                            imgs = drawing.image.Image('logo.png')
                            localOne.add_image(imgs)
                            localOne.sheet_view.zoomScale = 65
                            #wb.save(depense_Paquet)
                        self.lineNumber += 1
                        sheet= wb[row[headers[0]]]
                        j,m,a = headers[head].split()
                        monthNum = months.index(m)

                        if row[headers[0]] in numOccur :

                            occurName = numOccur.count(row[headers[0]])
                            s += occurName
                        else:
                            s = 13

                        sheet.cell(row=4,column=11).value = row[headers[0]]

                        sheet.cell(row=s,column=2).value = int(j)
                        sheet.cell(row=s,column=3).value = monthNum + 1
                        sheet.cell(row=s,column=4).value = int(a)
                        sheet.cell(row=s,column=6).value = self.foodPrice["Prov"]
                        sheet.cell(row=s,column=7).value = 'Mileage'

                        sheet.cell(row=s,column=14).value = int(self.foodPrice["Déjeuner"])
                        sheet.cell(row=s,column=15).value = int(self.foodPrice["Dîner"])
                        sheet.cell(row=s,column=16).value = int(self.foodPrice["Souper"])

                        sheet.cell(row=s,column=13).value = int(self.foodPrice["CV"])

                        numOccur.append(row[headers[0]])
                    # wb.save(depense_Paquet)



            except:
                Warning(data[headers[head]])

        self.personNumber  = len(wb.sheetnames) - 2

        try:
            #sheetToRemove = wb['Temp']
            #wb.remove_sheet(sheetToRemove)
            
            depenseSheets = wb.sheetnames
            annexeSheet = wb['Annexe']
            l = 10
            for sheett in depenseSheets[2:]:
                
                annexeSheet.cell(row=l,column=1).value  =  sheett
                annexeSheet.cell(row=l,column=2).value  = sheett
                annexeSheet.cell(row=l,column=4).value  = "='" + sheett + "'" + "!" + "X216"
                 # frm = "=SUMIF(" + sheett + "!" + "R:R;" + ">0;" + sheett +  "!X:X)"
                annexeSheet.cell(row=l,column=5).value  = "='" + sheett + "'" + "!" + "X218"
                annexeSheet.cell(row=l,column=6).value  = "='" + sheett + "'" + "!" + "X220"
                l += 1
            wb.save(depense_Paquet)
            
        except:
            Warning("Veuillez le fichier a traiter")


        
        